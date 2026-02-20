"""
ORAEX PSU Manager — Qualys Importer
"""
import os
import openpyxl
from database import get_connection, init_db

def import_qualys_scan(file_path, source_type):
    """
    Import Qualys scan reports into the database.
    source_type indicates the origin, e.g., 'GetNet' or 'PagoNxt'.
    """
    if not os.path.exists(file_path):
        print(f"❌ Qualys file not found: {file_path}")
        return False

    print(f"\\n📂 Opening Qualys Scan ({source_type}): {os.path.basename(file_path)}")
    print(f"⏳ Loading workbook...")

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    print(f"✅ Loaded {len(wb.sheetnames)} sheets")
    
    conn = get_connection()
    cursor = conn.cursor()
    
    total_detections = 0
    new_qids = 0
    
    try:
        if source_type == 'PagoNxt':
            count_det, count_qid = _import_pagonxt(wb, cursor)
        else:
            count_det, count_qid = _import_getnet(wb, cursor)
            
        total_detections += count_det
        new_qids += count_qid
        
        conn.commit()
        
        print(f"\\n{'='*50}")
        print(f"🎉 Qualys {source_type} import complete!")
        print(f"🛡️ New QIDs added: {new_qids}")
        print(f"🎯 Detections imported: {total_detections}")
        print(f"{'='*50}")
        
        return True
        
    except Exception as e:
        print(f"\\n❌ Qualys import error: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        conn.close()
        wb.close()

def _upsert_vulnerability(cursor, data):
    """Upsert na tabela de vulnerabilidades (catálogo global)."""
    try:
        qid = int(data['qid']) if data['qid'] else 0
        if not qid:
            return
    except:
        return

    # Verificar se já existe
    row = cursor.execute("SELECT qid FROM qualys_vulnerabilities WHERE qid = ?", (qid,)).fetchone()
    if not row:
        cursor.execute("""
            INSERT INTO qualys_vulnerabilities (qid, title, severity, threat, solution, category)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            qid,
            str(data.get('title', '')),
            str(data.get('severity', '')),
            str(data.get('threat', '')),
            str(data.get('solution', '')),
            str(data.get('category', ''))
        ))

def _insert_detection(cursor, data):
    """Insert na tabela de detecções por servidor."""
    try:
        qid = int(data['qid']) if data['qid'] else 0
        if not qid or not data['asset_name']:
            return 0
    except:
        return 0
    
    cursor.execute("""
        INSERT INTO qualys_detections (
            qid, asset_name, asset_ip, environment, os, os_version, 
            status, first_detected, last_detected, detection_age, 
            results, overdue, source
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        qid,
        str(data.get('asset_name', '')),
        str(data.get('asset_ip', '')),
        str(data.get('environment', '')),
        str(data.get('os', '')),
        str(data.get('os_version', '')),
        str(data.get('status', 'Active')),
        str(data.get('first_detected', '')),
        str(data.get('last_detected', '')),
        data.get('detection_age', 0),
        str(data.get('results', '')),
        str(data.get('overdue', '')),
        data.get('source')
    ))
    return 1

def _import_pagonxt(wb, cursor):
    """Parse PagoNxt specific Qualys format (sheet: DEMANDAS PM)"""
    sheet_name = 'DEMANDAS PM'
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found in PagoNxt spreadsheet.")
        return 0, 0
        
    ws = wb[sheet_name]
    count_det = 0
    count_qid = 0
    
    # Índices com base no qualys_headers.json
    # Asset Name: 0, Title: 1, Results: 2, Detection AGE: 3, First Detected: 4
    # Ambiente: 6, OS: 7, OS Vers.: 8, Severity: 12, Last Detected: 13, 
    # Asset IPV4: 14, Solution: 15, QID: 16, Overdue: 17
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or row[0] == 'Asset Name':
            continue
            
        data = {
            'asset_name': row[0],
            'title': row[1],
            'results': row[2],
            'detection_age': row[3] if row[3] else 0,
            'first_detected': row[4],
            'environment': row[6],
            'os': row[7],
            'os_version': row[8],
            'severity': row[12],
            'last_detected': row[13],
            'asset_ip': row[14],
            'solution': row[15],
            'qid': row[16],
            'overdue': row[17],
            'source': 'PagoNxt'
        }
        
        # Oupsert Vulnerability
        _upsert_vulnerability(cursor, data)
        count_qid += 1  # Not exact unique count, but close enough for logs
        
        # Insert Detection
        count_det += _insert_detection(cursor, data)
        
    return count_det, count_qid

def _import_getnet(wb, cursor):
    """Parse GetNet specific Qualys format (sheet: PROCV)"""
    sheet_name = 'PROCV'
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found in GetNet spreadsheet.")
        return 0, 0
        
    ws = wb[sheet_name]
    count_det = 0
    count_qid = 0
    
    # Índices (PROCV na col 0)
    # Asset Name: 1, Title: 2, Results: 3, Detection AGE: 4, First Detected: 5
    # Ambiente: 8, Sistema Operacional: 9, Versão de SO: 10
    # Severity: 14, Last Detected: 15, Asset IPV4: 16, Solution: 17, QID: 18, Overdue: 19
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or row[1] == 'Asset Name' or row[1] == '#N/D':
            # Em planilhas grandes PROCV gera #N/D quando não encontra
            continue
            
        data = {
            'asset_name': row[1],
            'title': row[2],
            'results': row[3],
            'detection_age': row[4] if row[4] else 0,
            'first_detected': row[5],
            'environment': row[8],
            'os': row[9],
            'os_version': row[10],
            'severity': row[14],
            'last_detected': row[15],
            'asset_ip': row[16],
            'solution': row[17],
            'qid': row[18],
            'overdue': row[19],
            'source': 'GetNet'
        }
        
        _upsert_vulnerability(cursor, data)
        count_det += _insert_detection(cursor, data)
        
    # Retorna o total aproximado de novas vulns
    count_qid = cursor.execute("SELECT COUNT(*) FROM qualys_vulnerabilities").fetchone()[0]
    
    return count_det, count_qid

if __name__ == "__main__":
    init_db()
    # Test script locally with hardcoded paths
    pagonxt_path = r'd:\\antigravity\\oraex-psu\\scan-vulnerabilidades\\20260219 - SCAN FULL QUALYS - PAGONXT.xlsx'
    getnet_path = r'd:\\antigravity\\oraex-psu\\scan-vulnerabilidades\\20260219 - SCAN FULL QUALYS.xlsm'
    
    import_qualys_scan(pagonxt_path, 'PagoNxt')
    import_qualys_scan(getnet_path, 'GetNet')
