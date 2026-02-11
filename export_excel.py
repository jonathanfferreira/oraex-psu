import shutil
import os
import openpyxl
from datetime import datetime
import locale
from config import GMUD_PATH

# Tentar configurar locale para PT-BR para nomes de dias da semana
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
    except:
        pass

# Diret√≥rio para backups autom√°ticos antes de cada escrita
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')

if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

def backup_workbook(file_path):
    """
    Cria uma c√≥pia de seguran√ßa do arquivo Excel antes de qualquer modifica√ß√£o.
    Salva em ./backups/ com timestamp.
    """
    try:
        if not os.path.exists(file_path):
            return False, "Arquivo original n√£o encontrado para backup."

        filename = os.path.basename(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"{filename}_{timestamp}.bak")
        
        shutil.copy2(file_path, backup_path)
        print(f"üì¶ Backup criado em: {backup_path}")
        return True, backup_path
    except Exception as e:
        return False, str(e)

def get_sheet_name_for_date(date_obj):
    """
    Retorna o nome da aba correspondente ao M√™s/Ano da data fornecida.
    Baseado no mapeamento do config.py (Reverso).
    """
    from config import MONTH_SHEETS
    target_year = date_obj.year
    target_month = date_obj.month
    
    for sheet_name, (year, month) in MONTH_SHEETS.items():
        if year == target_year and month == target_month:
            return sheet_name
            
    return None

def write_gmud_to_excel(gmud_data):
    """
    Escreve os dados do GMUD na planilha Excel original.
    
    Args:
        gmud_data (dict): Dicion√°rio contendo os dados do formul√°rio.
    
    Returns:
        tuple: (sucesso: bool, mensagem: str)
    """
    file_path = GMUD_PATH
    
    # 1. Realizar Backup
    ok, msg = backup_workbook(file_path)
    if not ok:
        return False, f"Falha ao criar backup: {msg}"

    try:
        # 2. Carregar Workbook (keep_vba=True para preservar macros)
        # data_only=False para manter f√≥rmulas, keep_vba=True para macros
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        
        # 3. Identificar a Aba Correta
        start_date_str = gmud_data.get('start_date')
        if isinstance(start_date_str, str):
            # Formato esperado do frontend: YYYY-MM-DDTHH:MM
            start_date = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
        else:
            start_date = start_date_str

        end_date_str = gmud_data.get('end_date')
        if isinstance(end_date_str, str):
            end_date = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
        else:
            end_date = end_date_str
            
        sheet_name = get_sheet_name_for_date(start_date)
        if not sheet_name or sheet_name not in wb.sheetnames:
            possible_sheets = ", ".join(wb.sheetnames)
            return False, f"Aba para a data {start_date} n√£o encontrada. Abas dispon√≠veis: {possible_sheets}"
        
        ws = wb[sheet_name]
        
        # 4. Encontrar a pr√≥xima linha vazia
        # Come√ßa da linha 2 (assumindo cabe√ßalho na linha 1)
        row_idx = 2
        # Verifica coluna I (T√≠tulo - √≠ndice 9) ou coluna H (GMUD - √≠ndice 8) para saber se a linha est√° ocupada
        while ws.cell(row=row_idx, column=9).value or ws.cell(row=row_idx, column=8).value:
            row_idx += 1
            
        # 5. Escrever Dados nas Colunas (Mapeamento baseado em import_excel.py)
        # A (1): Cliente
        # B (2): Tipo BD
        # C (3): Entorno
        # D (4): Status
        # E (5): Dia
        # F (6): Data In√≠cio
        # G (7): Data T√©rmino
        # H (8): GMUD / Change Number
        # I (9): T√≠tulo
        # J (10): Designado a
        # K (11): Observa√ß√£o
        # L (12): Vulnerabilidade
        # M (13): Aberto Por
        
        ws.cell(row=row_idx, column=1, value=gmud_data.get('client', 'Getnet'))
        ws.cell(row=row_idx, column=2, value=gmud_data.get('db_type', 'Oracle'))
        ws.cell(row=row_idx, column=3, value=gmud_data.get('environment', 'PROD'))
        ws.cell(row=row_idx, column=4, value=gmud_data.get('status', 'Planejada'))
        
        # Dia da semana (em portugu√™s se locale funcionar, sen√£o ingl√™s)
        day_name = start_date.strftime('%A').capitalize()
        ws.cell(row=row_idx, column=5, value=day_name)
        
        ws.cell(row=row_idx, column=6, value=start_date)
        ws.cell(row=row_idx, column=7, value=end_date)
        
        ws.cell(row=row_idx, column=8, value=gmud_data.get('change_number', 'N/A'))
        ws.cell(row=row_idx, column=9, value=gmud_data.get('title'))
        ws.cell(row=row_idx, column=10, value=gmud_data.get('assigned_to'))
        ws.cell(row=row_idx, column=11, value=gmud_data.get('observation', ''))
        ws.cell(row=row_idx, column=12, value=gmud_data.get('vulnerability', ''))
        ws.cell(row=row_idx, column=13, value=gmud_data.get('opened_by', ''))

        # 6. Salvar Arquivo
        wb.save(file_path)
        print(f"‚úÖ GMUD gravada com sucesso na linha {row_idx} da aba {sheet_name}")
        
        return True, f"GMUD criada com sucesso na linha {row_idx}!"

    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"Erro ao escrever no Excel: {str(e)}"
