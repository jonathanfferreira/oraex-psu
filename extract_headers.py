import openpyxl
import os
import json

files = [
    r'd:\antigravity\oraex-psu\scan-vulnerabilidades\20260219 - SCAN FULL QUALYS - PAGONXT.xlsx',
    r'd:\antigravity\oraex-psu\scan-vulnerabilidades\20260219 - SCAN FULL QUALYS.xlsm'
]

results = {}

for f in files:
    filename = os.path.basename(f)
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        results[filename] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # find first row with data
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                non_empty = [str(x) for x in row if x is not None and str(x).strip() != '']
                if len(non_empty) > 3: # likely a header row
                    headers = non_empty
                    break
            results[filename][sheet_name] = headers
            print(f"File: {filename} | Sheet: {sheet_name} | Headers found: {len(headers)}")
    except Exception as e:
        results[filename] = f"Error: {str(e)}"

with open(r'd:\antigravity\oraex-psu\qualys_headers.json', 'w', encoding='utf-8') as outfile:
    json.dump(results, outfile, indent=2, ensure_ascii=False)
print("Saved headers to qualys_headers.json")
