"""
ORAEX PSU Manager — Excel Import Script
Reads the consolidation spreadsheet and imports data into PostgreSQL.
"""
import openpyxl
import psycopg2
import os
import sys
from datetime import datetime
from config import EXCEL_PATH, MONTH_SHEETS
from database import init_db, get_connection


def safe_str(value):
    """Convert cell value to string safely."""
    if value is None:
        return ""
    return str(value).strip()


def safe_datetime(value):
    """Convert cell value to datetime string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def normalize_gmud_status(status):
    """Normalize GMUD status values — maps emojis, typos, and free text to standard values."""
    if not status or not status.strip():
        return "NOVO"
    s = status.strip()
    mapping = {
        "\u2705": "ENCERRADA",
        "\U0001f504": "REPLANEJAR",
        "\U0001f504\U0001f4c5": "REPLANEJAR",
        "\U0001f6ab": "CANCELADA",
        "PROGAMADO": "PROGRAMADA",
        "AVALIAR\U0001f4c5": "AVALIAR",
        "NOVO\U0001f4c5": "NOVO",
        "AUTORIZAR\U0001f4c5": "AUTORIZAR",
        "re": "REPLANEJAR",
        "CANCELAR": "CANCELADA",
    }
    if s in mapping:
        return mapping[s]
    if s.lower().startswith("freezin"):
        return "FREEZING"
    return s


def normalize_cmdb_status(status):
    """Normalize CMDB Full status values — unify casing and remove REQ suffixes."""
    if not status or not status.strip():
        return ""
    s = status.strip()
    low = s.lower()
    if low.startswith("descontinuado"):
        return "Descontinuado"
    if low == "sendo descontinuado":
        return "Sendo Descontinuado"
    if low.startswith("stopped"):
        return "Descontinuado"
    return s


def import_servers(wb, conn):
    """Import from 'GetNet - Oracle Databases' sheet.
    
    Counts standby/contingency as a separate server (total_servers=2).
    Detects GGS flag '(G)' in hostnames.
    """
    sheet_name = "GetNet - Oracle Databases"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    cursor.execute("DELETE FROM servers")
    count = 0
    total_server_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]
        # Skip empty rows
        if not any(vals[:5]):
            continue

        primary = safe_str(vals[1])
        standby = safe_str(vals[2])

        # Detect if standby/contingency exists → counts as a second server
        has_standby = 1 if (standby and standby not in ('N/A', '', 'None')) else 0
        num_servers = 2 if has_standby else 1

        # Detect GGS: servers with "(G)" in hostname
        has_ggs = 1 if ('(G)' in primary or '(G)' in standby) else 0

        # Clean PSU version — "Descontinuado" is a status, not a version
        psu_version = safe_str(vals[3])
        if psu_version.lower() == "descontinuado":
            psu_version = ""

        cursor.execute("""
            INSERT INTO servers (environment, primary_hostname, standby_hostname, psu_version,
                email_sent, alignment, ggs_version, primary_contact, responsible_team,
                system_product, application_day, start_time, end_time, observation,
                total_servers, has_standby, has_ggs)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            safe_str(vals[0]),   # A: ENVIROMENT
            primary,             # B: PRIMARY HOSTNAME
            standby,             # C: STANDBY HOSTNAME
            psu_version,         # D: GRID/PSU VERSION (cleaned)
            safe_str(vals[4]),   # E: Email Enviado
            safe_str(vals[5]),   # F: Alinhamento
            safe_str(vals[6]),   # G: GGS VERSION
            safe_str(vals[7]),   # H: Contato Primário
            safe_str(vals[8]),   # I: Equipe Responsável
            safe_str(vals[9]),   # J: Sistema/Serviço/Produto
            safe_str(vals[10]),  # K: Dia para aplicação
            safe_str(vals[11]),  # L: Horário de Início
            safe_str(vals[12]),  # M: Horário de Fim
            safe_str(vals[13]),  # N: OBSERVAÇÃO
            num_servers,         # Total Servers (1 standalone, 2 with standby)
            has_standby,         # Has standby flag
            has_ggs,             # Has GGS flag
        ))
        count += 1
        total_server_count += num_servers

    conn.commit()
    print(f"   Servers: {count} rows imported ({total_server_count} total servers, including standby)")
    return count


def import_cmdb(wb, conn):
    """Import from 'GetNet CMDB - Databases' sheet."""
    sheet_name = "GetNet CMDB - Databases"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    cursor.execute("DELETE FROM cmdb_databases")
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]
        if not any(vals[:5]):
            continue

        cursor.execute("""
            INSERT INTO cmdb_databases (environment, name, contingency_name, db_type, db_version,
                db_version_detail, status, application_day, week_month, start_time, end_time,
                system, system_product, type, os, primary_contact, function, description,
                os_type, responsible_team, manager, team_email, validation_contact, ip_address)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            safe_str(vals[0]),   # A: Entorno
            safe_str(vals[1]),   # B: Nome
            safe_str(vals[2]),   # C: Nome da Contingência
            safe_str(vals[3]),   # D: Nome do Banco (db_type)
            safe_str(vals[4]),   # E: Versao Banco
            safe_str(vals[5]),   # F: DB Version
            safe_str(vals[6]),   # G: Situação
            safe_str(vals[7]),   # H: Dia para aplicação
            safe_str(vals[8]),   # I: Semanas (mês)
            safe_str(vals[9]),   # J: Horário de Início
            safe_str(vals[10]),  # K: Horário de Fim
            safe_str(vals[11]),  # L: Sistema
            safe_str(vals[12]),  # M: Sistema/Serviço/Produto
            safe_str(vals[13]),  # N: Tipo
            safe_str(vals[14]),  # O: Sistema Operacional
            safe_str(vals[15]),  # P: Contato Primário
            safe_str(vals[16]),  # Q: Função
            safe_str(vals[17]),  # R: Descrição do IC
            safe_str(vals[18]),  # S: Tipo SO
            safe_str(vals[19]),  # T: Equipe Responsável
            safe_str(vals[20]),  # U: Gerente
            safe_str(vals[21]),  # V: E-mail
            safe_str(vals[22]),  # W: Validação
            safe_str(vals[23]),  # X: IP
        ))
        count += 1

    conn.commit()
    print(f"   CMDB Databases: {count} records imported")
    return count


def import_gmuds(wb, conn):
    """Import from monthly sheets (FEVEREIRO-25 through FEVEREIRO-26)."""
    cursor = conn.cursor()
    cursor.execute("DELETE FROM gmuds")
    total_count = 0

    for sheet_name, (year, month) in MONTH_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        count = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            vals = [cell.value for cell in row]
            # Skip empty rows - need at least a title or change number
            if not any(vals[:10]):
                continue

            # Columns vary per month sheet, but base pattern (cols 1-13):
            # A: Cliente, B: Tipo BD, C: Entorno, D: Status, E: Dia,
            # F: Data Início, G: Data Término, H: GMUD, I: Título,
            # J: Designado a, K: Observação, L: Vulnerabilidade, M: Aberto Por
            #
            # Extended cols (Jul, Ago, Out, Nov):
            # L: Vulnerabilidade Antes, M: Vulnerabilidade Após,
            # N: Código fechamento, O: Replanejar?, P: Nova Data Início,
            # Q: Nova Data Fim, R: Nova GMUD

            # Detect extended layout: cols 12+ have "Antes/Após" pattern
            # In extended sheets, col 12 = Vuln Antes, col 13 = Vuln Após
            # In classic sheets, col 12 = Vulnerabilidade, col 13 = Aberto Por
            num_cols = len(vals)
            is_extended = num_cols > 13 and any(
                safe_str(vals[i]) for i in range(13, min(18, num_cols))
            )

            if is_extended:
                # Extended layout (Jul, Ago, Out, Nov)
                vulnerability = ""  # No single "vulnerability" field
                opened_by = ""     # No "opened_by" in extended layout
                vuln_before = safe_str(vals[11]) if num_cols > 11 else ""
                vuln_after = safe_str(vals[12]) if num_cols > 12 else ""
                closing_code = safe_str(vals[13]) if num_cols > 13 else ""
                needs_replan = safe_str(vals[14]) if num_cols > 14 else ""
                new_start = safe_datetime(vals[15]) if num_cols > 15 else ""
                new_end = safe_datetime(vals[16]) if num_cols > 16 else ""
                new_gmud = safe_str(vals[17]) if num_cols > 17 else ""
            else:
                # Classic layout (Fev-Jun, Dez, Jan, Fev-26)
                vulnerability = safe_str(vals[11]) if num_cols > 11 else ""
                opened_by = safe_str(vals[12]) if num_cols > 12 else ""
                vuln_before = ""
                vuln_after = ""
                closing_code = ""
                needs_replan = ""
                new_start = ""
                new_end = ""
                new_gmud = ""

            cursor.execute("""
                INSERT INTO gmuds (year, month, client, db_type, environment, status,
                    day_of_week, start_date, end_date, change_number, title,
                    assigned_to, observation, vulnerability, opened_by,
                    vulnerability_before, vulnerability_after, closing_code,
                    needs_replan, new_start_date, new_end_date, new_gmud)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                year,
                month,
                safe_str(vals[0]),   # A: Cliente
                safe_str(vals[1]),   # B: Tipo BD
                safe_str(vals[2]),   # C: Entorno
                normalize_gmud_status(safe_str(vals[3])),   # D: Status (normalized)
                safe_str(vals[4]),   # E: Dia
                safe_datetime(vals[5]),  # F: Data Início
                safe_datetime(vals[6]),  # G: Data Término
                safe_str(vals[7]),   # H: GMUD
                safe_str(vals[8]),   # I: Título
                safe_str(vals[9]),   # J: Designado a
                safe_str(vals[10]) if num_cols > 10 else "",  # K: Observação
                vulnerability,
                opened_by,
                vuln_before,
                vuln_after,
                closing_code,
                needs_replan,
                new_start,
                new_end,
                new_gmud,
            ))
            count += 1

        total_count += count
        print(f"   {sheet_name}: {count} GMUDs imported")

    conn.commit()
    print(f"   Total GMUDs: {total_count} records imported")
    return total_count


def import_planning(wb, conn):
    """Import from 'Planejamento oracle' sheet."""
    sheet_name = "Planejamento oracle"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    cursor.execute("DELETE FROM planning")
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]
        if not any(vals[:5]):
            continue

        cursor.execute("""
            INSERT INTO planning (hostname, contingency_name, application_day, week_month,
                start_time, end_time, primary_contact, db_version, bank_version,
                system, system_product, os, function, description, responsible_team,
                validation_contact)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            safe_str(vals[0]),   # A: Nome
            safe_str(vals[1]),   # B: Nome da Contingência
            safe_str(vals[2]),   # C: Dia para aplicação
            safe_str(vals[3]),   # D: Semanas (mês)
            safe_str(vals[4]),   # E: Horário de Início
            safe_str(vals[5]),   # F: Horário de Fim
            safe_str(vals[6]),   # G: Contato Primário
            safe_str(vals[7]),   # H: DB Version
            safe_str(vals[8]),   # I: Versão Banco
            safe_str(vals[9]),   # J: Sistema
            safe_str(vals[10]) if len(vals) > 10 else "",  # K: Sistema/Produto
            safe_str(vals[11]) if len(vals) > 11 else "",  # L: SO
            safe_str(vals[12]) if len(vals) > 12 else "",  # M: Função
            safe_str(vals[13]) if len(vals) > 13 else "",  # N: Descrição
            safe_str(vals[14]) if len(vals) > 14 else "",  # O: Equipe
            safe_str(vals[15]) if len(vals) > 15 else "",  # P: Validação
        ))
        count += 1

    conn.commit()
    print(f"   Planning: {count} records imported")
    return count


def import_pagonxt(wb, conn):
    """Import from 'PagoNxt - Databases' sheet."""
    sheet_name = "PagoNxt - Databases"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    cursor.execute("DELETE FROM pagonxt_databases")
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]
        if not any(vals[:5]):
            continue

        cursor.execute("""
            INSERT INTO pagonxt_databases (environment, name, contingent, psu_version,
                contact, zone, product, description, channel, service, observation,
                ip, instance, status, os)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            safe_str(vals[0]),   # A: ENVIROLMENT
            safe_str(vals[1]),   # B: NAME
            safe_str(vals[2]),   # C: CONTINGENT
            safe_str(vals[3]),   # D: GRID/PSU VERSION
            safe_str(vals[4]),   # E: CONTACT
            safe_str(vals[5]),   # F: ZONE
            safe_str(vals[6]),   # G: PRODUCT
            safe_str(vals[7]),   # H: DESCRIPTION
            safe_str(vals[8]),   # I: CHANNEL
            safe_str(vals[9]),   # J: SERVICE
            safe_str(vals[10]) if len(vals) > 10 else "",  # K: OBSERVATION
            safe_str(vals[11]) if len(vals) > 11 else "",  # L: IP
            safe_str(vals[12]) if len(vals) > 12 else "",  # M: INSTANCE
            safe_str(vals[13]) if len(vals) > 13 else "",  # N: STATUS
            safe_str(vals[14]) if len(vals) > 14 else "",  # O: OS
        ))
        count += 1

    conn.commit()
    print(f"   PagoNxt Databases: {count} records imported")
    return count


def run_import(excel_path=None):
    """Run the full import process."""
    path = excel_path or EXCEL_PATH

    if not os.path.exists(path):
        print(f" Excel file not found: {path}")
        return False

    print(f" Opening: {os.path.basename(path)}")
    print(f" Loading workbook...")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f" Loaded {len(wb.sheetnames)} sheets")

    # Initialize DB
    init_db()
    conn = get_connection()

    total = 0
    sheets_imported = []

    try:
        print("\n Importing data...\n")

        # Import each data source
        count = import_servers(wb, conn)
        total += count
        if count > 0:
            sheets_imported.append("servers")

        count = import_cmdb(wb, conn)
        total += count
        if count > 0:
            sheets_imported.append("cmdb")

        count = import_gmuds(wb, conn)
        total += count
        if count > 0:
            sheets_imported.append("gmuds")

        count = import_planning(wb, conn)
        total += count
        if count > 0:
            sheets_imported.append("planning")

        count = import_pagonxt(wb, conn)
        total += count
        if count > 0:
            sheets_imported.append("pagonxt")

        # Log the import
        _cur = conn.cursor()
        _cur.execute("""
            INSERT INTO import_log (source_file, sheets_imported, total_records, status, message)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            os.path.basename(path),
            ", ".join(sheets_imported),
            total,
            "success",
            f"Imported {total} total records from {len(sheets_imported)} data sources"
        ))
        conn.commit()

        print(f"\n{'='*50}")
        print(f" Import complete! {total} total records imported.")
        print(f" Sources: {', '.join(sheets_imported)}")
        print(f"{'='*50}")

        return True

    except Exception as e:
        print(f"\n Import error: {e}")
        _cur = conn.cursor()
        _cur.execute("""
            INSERT INTO import_log (source_file, sheets_imported, total_records, status, message)
            VALUES (%s, %s, %s, %s, %s)
        """, (os.path.basename(path), "", 0, "error", str(e)))
        conn.commit()
        raise
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    run_import(excel_path)


# ══════════════════════════════════════════════════════════
#  CMDB FULL IMPORT (separate spreadsheet)
# ══════════════════════════════════════════════════════════

def import_cmdb_full_getnet(wb, conn):
    """Import DB servers from 'CMDB Geral GETNET Brasil'.
    Only rows where 'Banco de Dados' (col 11) is non-empty.
    """
    sheet_name = "CMDB Geral GETNET Brasil"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        db_type = safe_str(row[11]).strip() if row[11] else ""
        if not db_type or db_type in ("None", "N/A"):
            continue

        # Normalize common DB type names
        db_type = normalize_db_type(db_type)
        if not db_type:
            continue

        cursor.execute("""
            INSERT INTO cmdb_full (client, hostname, contingency, db_type, db_version,
                status, server_type, environment, os, responsible_team, manager,
                primary_contact, system_product, function, description,
                validation_contact, team_email, shutdown_procedure, affinity,
                week_month, application_day, start_time, end_time,
                importance_level, criticality, scope_pci, scope_sox, scope_pagonxt,
                ip_service, ip_backup, ip_branca, source_sheet)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            "GetNet",
            safe_str(row[8]),    # Nome
            safe_str(row[9]),    # Contingência
            db_type,             # Banco de Dados (normalized)
            safe_str(row[12]),   # DB Version
            normalize_cmdb_status(safe_str(row[18])),   # Situação (normalized)
            safe_str(row[19]),   # Tipo
            safe_str(row[20]),   # Ambiente
            safe_str(row[21]),   # Sistema Operacional
            safe_str(row[22]),   # Equipe Responsável
            safe_str(row[23]),   # Gerente
            safe_str(row[24]),   # Contato Primário
            safe_str(row[25]),   # Sistema/Serviço/Produto
            safe_str(row[26]),   # Função
            safe_str(row[27]),   # Descrição do IC
            safe_str(row[28]) if len(row) > 28 else "",  # Validação
            safe_str(row[29]) if len(row) > 29 else "",  # E-mail
            safe_str(row[30]) if len(row) > 30 else "",  # Procedimento
            safe_str(row[32]) if len(row) > 32 else "",  # Afinidade
            safe_str(row[33]) if len(row) > 33 else "",  # Semanas
            safe_str(row[34]) if len(row) > 34 else "",  # Dia para aplicação
            safe_str(row[35]) if len(row) > 35 else "",  # Horário Início
            safe_str(row[36]) if len(row) > 36 else "",  # Horário Fim
            safe_str(row[37]) if len(row) > 37 else "",  # Nível Importância
            safe_str(row[38]) if len(row) > 38 else "",  # Criticidade
            safe_str(row[39]) if len(row) > 39 else "",  # Escopo PCI
            safe_str(row[40]) if len(row) > 40 else "",  # Escopo SOX
            safe_str(row[41]) if len(row) > 41 else "",  # Escopo PagoNxt
            safe_str(row[53]) if len(row) > 53 else "",  # IP Serviço
            safe_str(row[54]) if len(row) > 54 else "",  # IP Backup
            safe_str(row[55]) if len(row) > 55 else "",  # IP Branca
            sheet_name,
        ))
        count += 1

    conn.commit()
    print(f"   CMDB GetNet Brasil: {count} DB servers imported")
    return count


def import_cmdb_full_latam(wb, conn):
    """Import DB servers from 'CMDB Geral LATAM' (PagoNxt).
    Only rows where 'Banco de Dados' (col 11) is non-empty.
    """
    sheet_name = "CMDB Geral LATAM"
    if sheet_name not in wb.sheetnames:
        print(f"  ️  Sheet '{sheet_name}' not found, skipping.")
        return 0

    ws = wb[sheet_name]
    cursor = conn.cursor()
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        db_type = safe_str(row[11]).strip() if row[11] else ""
        if not db_type or db_type in ("None", "N/A"):
            continue

        db_type = normalize_db_type(db_type)
        if not db_type:
            continue

        cursor.execute("""
            INSERT INTO cmdb_full (client, hostname, contingency, db_type, db_version,
                status, server_type, environment, os, responsible_team, manager,
                primary_contact, system_product, function, description,
                validation_contact, team_email, shutdown_procedure, affinity,
                week_month, application_day, start_time, end_time,
                importance_level, criticality, scope_pci, scope_sox, scope_pagonxt,
                ip_service, ip_backup, ip_branca, zone, country, source_sheet)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            "PagoNxt",
            safe_str(row[8]),    # Nome
            safe_str(row[9]),    # Contingência
            db_type,             # Banco de Dados
            safe_str(row[12]),   # DB Version
            normalize_cmdb_status(safe_str(row[20])),   # Situação (normalized)
            safe_str(row[21]),   # Tipo
            safe_str(row[22]),   # Ambiente
            safe_str(row[23]),   # Sistema Operacional
            safe_str(row[24]),   # Equipe Responsável
            safe_str(row[25]),   # Gerente
            safe_str(row[26]),   # Contato Primário
            safe_str(row[27]),   # Sistema/Serviço/Produto
            safe_str(row[28]) if len(row) > 28 else "",  # Função
            safe_str(row[29]) if len(row) > 29 else "",  # Descrição
            safe_str(row[30]) if len(row) > 30 else "",  # Validação
            safe_str(row[31]) if len(row) > 31 else "",  # E-mail
            safe_str(row[32]) if len(row) > 32 else "",  # Procedimento
            safe_str(row[34]) if len(row) > 34 else "",  # Afinidade
            safe_str(row[35]) if len(row) > 35 else "",  # Semanas
            safe_str(row[36]) if len(row) > 36 else "",  # Dia para aplicação
            safe_str(row[37]) if len(row) > 37 else "",  # Horário Início
            safe_str(row[38]) if len(row) > 38 else "",  # Horário Fim
            safe_str(row[39]) if len(row) > 39 else "",  # Nível Importância
            safe_str(row[40]) if len(row) > 40 else "",  # Criticidade
            safe_str(row[41]) if len(row) > 41 else "",  # Escopo PCI
            safe_str(row[42]) if len(row) > 42 else "",  # Escopo SOX
            safe_str(row[43]) if len(row) > 43 else "",  # Escopo PagoNxt
            safe_str(row[55]) if len(row) > 55 else "",  # IP Serviço
            safe_str(row[56]) if len(row) > 56 else "",  # IP Backup
            safe_str(row[57]) if len(row) > 57 else "",  # IP Branca
            safe_str(row[18]) if len(row) > 18 else "",  # Zona
            safe_str(row[19]) if len(row) > 19 else "",  # País
            sheet_name,
        ))
        count += 1

    conn.commit()
    print(f"   CMDB LATAM (PagoNxt): {count} DB servers imported")
    return count


def normalize_db_type(db_type):
    """Normalize DB type names for consistent grouping."""
    db_lower = db_type.lower().strip()
    # Invalid values that should be skipped
    if db_lower in ("sim", "não", "nao", "no", "yes", "n/a", "-"):
        return ""
    mapping = {
        "oracle": "Oracle",
        "futuramente oracle": "Oracle",
        "golden gate(necessário stop)": "Oracle",
        "mongodb": "MongoDB",
        "mongodb (read)": "MongoDB",
        "mongodb (s)": "MongoDB",
        "mongodb (p)": "MongoDB",
        "mongo": "MongoDB",
        "sql server": "SQL Server",
        "sqlserver": "SQL Server",
        "mysql": "MySQL",
        "mariadb": "MariaDB",
        "postgres": "PostgreSQL",
        "postgresql": "PostgreSQL",
        "sybase": "Sybase",
        "sqlite3": "SQLite",
    }
    return mapping.get(db_lower, db_type)


def run_cmdb_full_import(excel_path=None):
    """Import CMDB Full spreadsheet (only DB servers)."""
    from config import CMDB_FULL_PATH

    path = excel_path or CMDB_FULL_PATH
    if not os.path.exists(path):
        print(f" CMDB Full file not found: {path}")
        return False

    print(f"\n Opening CMDB Full: {os.path.basename(path)}")
    print(f" Loading workbook...")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f" Loaded {len(wb.sheetnames)} sheets")

    init_db()
    conn = get_connection()

    # Clear previous CMDB Full data
    _cur = conn.cursor()
    _cur.execute("DELETE FROM cmdb_full")
    conn.commit()

    total = 0
    try:
        total += import_cmdb_full_getnet(wb, conn)
        total += import_cmdb_full_latam(wb, conn)

        print(f"\n{'='*50}")
        print(f" CMDB Full import complete! {total} DB servers imported.")
        print(f"{'='*50}")

        return True
    except Exception as e:
        print(f"\n CMDB Full import error: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        conn.close()
        wb.close()
